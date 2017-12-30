#!/usr/bin/env python3

from os import path, walk
from fnmatch import fnmatch
from datetime import datetime
from openpyxl import load_workbook

import pprint
import argparse
import logging as log

ROOT = "/home/mzaborow/Documents/Rekrutacje/jlabs"
FILE_PATTERN = "*_arkusz.xlsx"
MONTHLY_PATTERN = "????_??_m_zaborowski_raport.xlsx"

def iterate_over_interview_files():
    """Iterates over files - starts from ROOT, and goes over all subdirectories.
    Only directories with no subdirectories, are taken, and file name must follow FILE_PATTERN"""
    for dir, subdirs, files in walk(ROOT):
        if subdirs:
            log.debug("non empty directory - " + dir + " - ignoring")
            continue
        for name in files:
            if fnmatch(name, FILE_PATTERN):
                yield path.join(dir, name)
            else:
                log.debug("filename - " + name + " - not matching pattern (" + FILE_PATTERN + ") - ignoring")


def get_version(ws):
    """takes worksheet and finds last row in first column, there is version of test sheet there"""
    idx = 200
    while ws['A' + str(idx)].value in [None, '', 'None']:
        idx -=1
    return ws['A' + str(idx)].value


def is_valid(data):
    """Checks if passed data is valid. In practice it means that all entries are filled"""
    if data is None or not isinstance(data, dict):
        return False
    for buf in data.values():
        if buf in [None, '', 'None']:
            return False
    known = ['date', 'name', 'avg', 'base', 'java', 'frameworks', 'version']
    return len(set(known).intersection(data.keys())) == len(known)
    #TODO check elements' types


def extract_data_for_16_05_2017(ws):
    log.debug("extracting with format of 16.05.2017")
    buf = ws['B2'].value
    if buf is None:
        log.info("format of 16.05.2017 not valid")
        return None  # different format
    if not isinstance(buf, datetime):
        try:
            buf = datetime.strptime(ws['B2'].value, '%Y.%m.%d')
        except ValueError:
            buf = datetime.strptime(ws['B2'].value, '%d.%m.%Y')
    return {'date': buf,
            'name': ws['B4'].value,
            'avg': ws['B9'].value,
            'base': ws['B11'].value,
            'java': ws['B21'].value,
            'frameworks': ws['B29'].value,
            'version': get_version(ws)}


def extract_data_for_01_03_2017(ws):
    log.debug("extracting with format of 01.03.2017")
    buf = ws['B2'].value
    if buf is None:
        log.debug("of 01.03.2017 not valid")
        return None  # different format
    if not isinstance(buf, datetime):
        try:
            buf = datetime.strptime(ws['B2'].value, '%Y.%m.%d')
        except ValueError:
            buf = datetime.strptime(ws['B2'].value, '%d.%m.%Y')
    return {'date': buf,
            'name': ws['B4'].value,
            'avg': ws['B10'].value,
            'base': ws['B12'].value,
            'java': ws['B22'].value,
            'frameworks': ws['B30'].value,
            'version': get_version(ws)}


def extract_data_for_very_old(ws):
    log.debug("extracting with very old format")
    if not get_version(ws) == 'Uwagi (inne informacje warte odnotowania)':
        log.debug("very old format not valid - wrong version")
        return None
    buf = ws['B2'].value
    if buf is None:
        log.debug("very old format not valid - wrong interview date")
        return None  # old format
    if not isinstance(buf, datetime):
        try:
            buf = datetime.strptime(ws['B2'].value, '%Y.%m.%d')
        except ValueError:
            buf = datetime.strptime(ws['B2'].value, '%d.%m.%Y')
    return {'date': buf,
            'name': ws['B4'].value,
            'avg': ws['B10'].value,
            'base': ws['B12'].value,
            'java': ws['B23'].value,
            'frameworks': ws['B33'].value,
            'version': get_version(ws)}


def extract_interview_data(file_name):
    """From file, based on passed file name, extracts data - as dictionary"""
    wb = load_workbook(filename=file_name, data_only=True)
    ws = wb.active

    res = extract_data_for_16_05_2017(ws)
    if not is_valid(res):
        res = extract_data_for_01_03_2017(ws)
    if not is_valid(res):
        res = extract_data_for_very_old(ws)
    return res


def iterate_over_monthly_files():
    """Iterates over files - starts from ROOT, and goes over all subdirectories.
    Only directories with no subdirectories, are taken, and file name must follow FILE_PATTERN"""
    for dir, subdirs, files in walk(ROOT):
        for name in files:
            if fnmatch(name, MONTHLY_PATTERN):
                yield path.join(dir, name)
            else:
                log.debug("filename - " + name + " - not matching pattern (" + MONTHLY_PATTERN + ") - ignoring")


def extract_monthly_data(file_name):
    log.debug("preparing monthly report")
    log.debug(file_name)
    wb = load_workbook(filename=file_name, data_only=True)
    ws = wb.active
    res = {
        "name": file_name,
        "lines": []
    }
    idx = 2
    while not ws['B' + str(idx)].value in [None, '', 'None']:
        ln = {'no': ws['A' + str(idx)].value,
              'date': ws['B' + str(idx)].value,
              'recruter': ws['C' + str(idx)].value,
              'candidate_fname': ws['D' + str(idx)].value,
              'candidate_lname': ws['E' + str(idx)].value,
              'technology': ws['F' + str(idx)].value,
              'type': ws['G' + str(idx)].value,
              'avg': ws['H' + str(idx)].value,
              'base': ws['I' + str(idx)].value,
              'java': ws['J' + str(idx)].value,
              'frameworks': ws['K' + str(idx)].value}
        res['lines'].append(ln)
        idx += 1
    return res


def update_reports():
    pass


pp = pprint.PrettyPrinter(indent=2)

def processInterviewData(operation, dryRun):
    log.debug("processing interview data")
    # extract data
    entries = [extract_interview_data(file_name) for file_name in iterate_over_interview_files()]
    # filter out empty entries
    entries = [en for en in entries if not en is None]
    # filter for current month, and sort by date
    now = datetime.now()
    entries = [e for e in entries if e['date'].year == now.year ]
    entries = sorted(entries, key=lambda e: e['date'])
    # printout sample report
    for buf in entries:
        try:
            print("{date:%Y.%m.%d} {name:20s} {avg: 2.2f} {base: 2.2f} {java: 2.2f} {frameworks: 2.2f}".format(**buf), flush=True)
        except ValueError:
            print(buf)


def processMonthlyData(operation, dryRun):
    # extract data
    entries = [extract_monthly_data(file_name) for file_name in iterate_over_monthly_files()]
    print (entries)
    # filter out empty entries
    entries = [en for en in entries if not en is None]
    # pp.pprint(entries)
    ##data = prepare_monthy_reports(entries)
    entries = sorted(entries, key=lambda dt: dt['name'])
    for buf in entries:
        print("\n" + buf['name'])
        for line in buf['lines']:
            # print(line)
            print("  {no:02d} {date:%Y.%m.%d} {recruter:20s} {candidate_fname:10s} {candidate_lname:12s} - {technology} {type:4s}: {avg: 2.2f} {base: 2.2f} {java: 2.2f} {frameworks: 2.2f}".format(**line), flush=True)


def parseArgs():
    """ parse CLI input, and returns tuple of passed parameters """
    parser = argparse.ArgumentParser(
        __file__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
        description="Script used to operate on reports, aggregated reports.")
    parser.add_argument("-r", "--reports", nargs="+", choices=['interview', 'monthly'], required=True,
                        help="reporting level")
    parser.add_argument("-o", "--operation", choices=['print', 'printout', 'aggregate', 'none'], default='no_change',
                        help="operation to perform, default is none")
    parser.add_argument("-d", "--debug", action='store_const', const=True,
                        help="all debug stuff - on screen")
    parser.add_argument("-n", "--dry-run", action='store_const', const=True,
                        help="do not do anything - just printout what should be executed")
    args = parser.parse_args()
    return args.reports, args.operation, args.debug, args.dry_run

def main():
    reports, operation, debug, dryRun = parseArgs()

    if debug:
        log.basicConfig(format="%(levelname)s: %(message)s", level=log.DEBUG)
        log.info("started.")
    else:
        log.basicConfig(format="%(levelname)s: %(message)s")

    print(" processing " + str(reports))
    print(" operation is " + operation)
    if dryRun:
        print(" this is DRY-RUN nothing will be changed!")

    if 'interview' in reports:
        processInterviewData(operation, dryRun)
    if 'monthly' in reports:
        processMonthlyData(operation, dryRun)

if __name__ == "__main__":
    main()
